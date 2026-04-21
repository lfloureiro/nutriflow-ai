from __future__ import annotations

import argparse
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook


DEFAULT_BASE_URL = "http://127.0.0.1:8000"
DEFAULT_FILE = Path("data/imports/familia/CalendarioDeRefeicoes.xlsm")


RECIPE_CATEGORY_MAP = {
    "carne": "carne",
    "peixe": "peixe",
    "vegetariano": "vegetariano_leguminosas",
    "vegetariana": "vegetariano_leguminosas",
    "vegetariano/leguminosas": "vegetariano_leguminosas",
    "vegetariano / leguminosas": "vegetariano_leguminosas",
    "leguminosas": "vegetariano_leguminosas",
    "outra": "outra",
}

PROTEIN_MAP = {
    "frango": "frango",
    "vaca": "vaca",
    "vitela": "vaca",
    "carne de vaca": "vaca",
    "porco": "porco",
    "peru": "peru",
    "enchidos": "enchidos_processados",
    "enchidos/processados": "enchidos_processados",
    "processados": "enchidos_processados",
    "peixe": "peixe",
    "bacalhau": "peixe",
    "atum": "peixe",
    "salmao": "peixe",
    "salmão": "peixe",
    "ovos": "ovos",
    "ovo": "ovos",
    "leguminosas": "leguminosas",
    "grao": "leguminosas",
    "grão": "leguminosas",
    "feijao": "leguminosas",
    "feijão": "leguminosas",
    "lentilhas": "leguminosas",
    "queijo": "queijo_lacticinios",
    "lacticinios": "queijo_lacticinios",
    "lacticínios": "queijo_lacticinios",
    "feta": "queijo_lacticinios",
}

MEAL_SUITABILITY_MAP = {
    "almoco": "almoco",
    "almoço": "almoco",
    "jantar": "jantar",
    "ambos": "ambos",
    "tanto": "ambos",
}


@dataclass
class RecipeRow:
    name: str
    protein_raw: str | None
    side_raw: str | None
    other_raw: str | None
    group_raw: str | None


@dataclass
class IngredientRow:
    recipe_name: str
    ingredient_name: str
    quantity: str | None
    unit: str | None


class ApiClient:
    def __init__(self, base_url: str, timeout: int = 30) -> None:
        self.base_url = base_url.rstrip("/")
        self.timeout = timeout
        self.session = requests.Session()

    def _url(self, path: str) -> str:
        return f"{self.base_url}{path}"

    def get(self, path: str) -> Any:
        response = self.session.get(self._url(path), timeout=self.timeout)
        self._raise_for_status(response)
        return response.json()

    def post(self, path: str, payload: dict[str, Any]) -> Any:
        response = self.session.post(
            self._url(path),
            json=payload,
            timeout=self.timeout,
        )
        self._raise_for_status(response)
        return response.json()

    def patch(self, path: str, payload: dict[str, Any]) -> Any:
        response = self.session.patch(
            self._url(path),
            json=payload,
            timeout=self.timeout,
        )
        self._raise_for_status(response)
        return response.json()

    def delete(self, path: str) -> Any:
        response = self.session.delete(self._url(path), timeout=self.timeout)
        self._raise_for_status(response)
        if response.content:
            return response.json()
        return None

    @staticmethod
    def _raise_for_status(response: requests.Response) -> None:
        if response.ok:
            return

        try:
            payload = response.json()
            detail = payload.get("detail")
        except Exception:
            detail = response.text

        raise RuntimeError(f"{response.status_code} {response.request.method} {response.url} -> {detail}")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_key(value: Any) -> str:
    text = normalize_text(value).casefold()
    return "".join(
        char for char in unicodedata.normalize("NFKD", text) if not unicodedata.combining(char)
    ).strip()


def infer_recipe_category(group_raw: str | None, protein_raw: str | None, name: str) -> str | None:
    group_key = normalize_key(group_raw)
    protein_key = normalize_key(protein_raw)
    name_key = normalize_key(name)

    if group_key in RECIPE_CATEGORY_MAP:
        return RECIPE_CATEGORY_MAP[group_key]

    if protein_key in {"atum", "bacalhau", "peixe", "salmao"}:
        return "peixe"

    if protein_key in {"frango", "vaca", "vitela", "porco", "peru", "enchidos"}:
        return "carne"

    if protein_key in {"ovos", "ovo", "leguminosas", "grao", "feijao", "lentilhas", "queijo", "feta"}:
        return "vegetariano_leguminosas"

    if any(token in name_key for token in ["lentilhas", "legumes", "vegetar", "grao", "feijao"]):
        return "vegetariano_leguminosas"

    return None


def infer_protein(protein_raw: str | None, name: str) -> str | None:
    protein_key = normalize_key(protein_raw)
    name_key = normalize_key(name)

    if protein_key in PROTEIN_MAP:
        return PROTEIN_MAP[protein_key]

    for token, mapped in PROTEIN_MAP.items():
        if token in name_key:
            return mapped

    return None


def infer_meal_suitability(name: str, side_raw: str | None, other_raw: str | None) -> str:
    haystack = " ".join([normalize_key(name), normalize_key(side_raw), normalize_key(other_raw)])

    if "sopa" in haystack or "quiche" in haystack:
        return "jantar"

    if "salada" in haystack:
        return "almoco"

    return "ambos"


def build_recipe_description(row: RecipeRow) -> str | None:
    parts: list[str] = []

    if normalize_text(row.protein_raw):
        parts.append(f"Proteína: {normalize_text(row.protein_raw)}")
    if normalize_text(row.side_raw):
        parts.append(f"Acompanhamento: {normalize_text(row.side_raw)}")
    if normalize_text(row.other_raw):
        parts.append(f"Outros: {normalize_text(row.other_raw)}")

    if not parts:
        return None

    return " | ".join(parts)


def find_header_map(worksheet, required_headers: list[str]) -> tuple[int, dict[str, int]]:
    for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        header_map: dict[str, int] = {}
        for idx, value in enumerate(row):
            header_map[normalize_key(value)] = idx

        if all(normalize_key(header) in header_map for header in required_headers):
            return row_idx, header_map

    raise RuntimeError(f"Não foi possível encontrar cabeçalhos: {required_headers}")


def parse_recipes(workbook) -> list[RecipeRow]:
    ws = workbook["Receitas"]
    _, headers = find_header_map(
        ws,
        ["Refeição", "Proteina", "Acompanha", "Outros", "Grupo"],
    )

    rows: list[RecipeRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = normalize_text(row[headers[normalize_key("Refeição")]])
        if not name:
            continue

        rows.append(
            RecipeRow(
                name=name,
                protein_raw=normalize_text(row[headers[normalize_key("Proteina")]]) or None,
                side_raw=normalize_text(row[headers[normalize_key("Acompanha")]]) or None,
                other_raw=normalize_text(row[headers[normalize_key("Outros")]]) or None,
                group_raw=normalize_text(row[headers[normalize_key("Grupo")]]) or None,
            )
        )

    return rows


def parse_products(workbook) -> dict[str, str]:
    ws = workbook["Produtos"]
    _, headers = find_header_map(ws, ["Mantimento"])

    unit_index = headers.get(normalize_key("Unidade"))

    result: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = normalize_text(row[headers[normalize_key("Mantimento")]])
        if not name:
            continue

        unit = normalize_text(row[unit_index]) if unit_index is not None else ""
        result[name] = unit

    return result


def parse_recipe_ingredients(workbook) -> list[IngredientRow]:
    ws = workbook["Ingredientes por receita"]
    _, headers = find_header_map(
        ws,
        ["Receita", "Quantidade", "Medida", "Mantimento"],
    )

    rows: list[IngredientRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        recipe_name = normalize_text(row[headers[normalize_key("Receita")]])
        ingredient_name = normalize_text(row[headers[normalize_key("Mantimento")]])

        if not recipe_name or not ingredient_name:
            continue

        quantity = normalize_text(row[headers[normalize_key("Quantidade")]]) or None
        unit = normalize_text(row[headers[normalize_key("Medida")]]) or None

        rows.append(
            IngredientRow(
                recipe_name=recipe_name,
                ingredient_name=ingredient_name,
                quantity=quantity,
                unit=unit,
            )
        )

    return rows


def group_ingredients_by_recipe(rows: list[IngredientRow]) -> dict[str, list[IngredientRow]]:
    grouped: dict[str, list[IngredientRow]] = {}
    for row in rows:
        grouped.setdefault(row.recipe_name, []).append(row)
    return grouped


def reset_catalog(api: ApiClient) -> None:
    print("==> Modo reset: a tentar apagar receitas existentes")
    recipes = api.get("/recipes/")
    for recipe in sorted(recipes, key=lambda item: item["id"], reverse=True):
        try:
            api.delete(f"/recipes/{recipe['id']}")
            print(f"   Receita apagada: {recipe['name']}")
        except Exception as exc:
            print(f"   [AVISO] Não foi possível apagar receita '{recipe['name']}': {exc}")

    print("==> Modo reset: a tentar apagar ingredientes existentes")
    ingredients = api.get("/ingredients/")
    for ingredient in sorted(ingredients, key=lambda item: item["id"], reverse=True):
        try:
            api.delete(f"/ingredients/{ingredient['id']}")
            print(f"   Ingrediente apagado: {ingredient['name']}")
        except Exception as exc:
            print(f"   [AVISO] Não foi possível apagar ingrediente '{ingredient['name']}': {exc}")


def upsert_ingredients(
    api: ApiClient,
    product_units: dict[str, str],
) -> dict[str, int]:
    existing = api.get("/ingredients/")
    ingredient_by_key = {normalize_key(item["name"]): item for item in existing}

    created = 0
    reused = 0

    for product_name in sorted(product_units.keys(), key=normalize_key):
        key = normalize_key(product_name)
        if key in ingredient_by_key:
            reused += 1
            continue

        created_item = api.post("/ingredients/", {"name": product_name})
        ingredient_by_key[key] = created_item
        created += 1

    print(f"==> Ingredientes: {created} criados, {reused} reutilizados")
    return {key: int(value["id"]) for key, value in ingredient_by_key.items()}


def upsert_recipes(
    api: ApiClient,
    recipes: list[RecipeRow],
) -> dict[str, int]:
    existing = api.get("/recipes/")
    recipe_by_key = {normalize_key(item["name"]): item for item in existing}

    created = 0
    updated = 0

    for recipe in recipes:
        payload = {
            "name": recipe.name,
            "description": build_recipe_description(recipe),
            "categoria_alimentar": infer_recipe_category(recipe.group_raw, recipe.protein_raw, recipe.name),
            "proteina_principal": infer_protein(recipe.protein_raw, recipe.name),
            "adequado_refeicao": infer_meal_suitability(recipe.name, recipe.side_raw, recipe.other_raw),
            "auto_plan_enabled": True,
        }

        key = normalize_key(recipe.name)

        if key in recipe_by_key:
            recipe_id = recipe_by_key[key]["id"]
            updated_item = api.patch(f"/recipes/{recipe_id}", payload)
            recipe_by_key[key] = updated_item
            updated += 1
        else:
            created_item = api.post("/recipes/", payload)
            recipe_by_key[key] = created_item
            created += 1

    print(f"==> Receitas: {created} criadas, {updated} atualizadas")
    return {key: int(value["id"]) for key, value in recipe_by_key.items()}


def replace_recipe_ingredients(
    api: ApiClient,
    recipe_ids_by_key: dict[str, int],
    ingredient_ids_by_key: dict[str, int],
    grouped_ingredients: dict[str, list[IngredientRow]],
) -> None:
    updated_recipes = 0

    for recipe_name, rows in grouped_ingredients.items():
        recipe_key = normalize_key(recipe_name)
        recipe_id = recipe_ids_by_key.get(recipe_key)

        if recipe_id is None:
            print(f"   [AVISO] Receita não encontrada para ingredientes: {recipe_name}")
            continue

        recipe_detail = api.get(f"/recipes/{recipe_id}")
        existing_links = recipe_detail.get("ingredient_links", [])

        for link in existing_links:
            api.delete(f"/recipes/{recipe_id}/ingredients/{link['id']}")

        seen_ingredient_keys: set[str] = set()

        for row in rows:
            ingredient_key = normalize_key(row.ingredient_name)
            ingredient_id = ingredient_ids_by_key.get(ingredient_key)

            if ingredient_id is None:
                print(f"   [AVISO] Ingrediente não encontrado: {row.ingredient_name}")
                continue

            duplicate_key = f"{recipe_id}:{ingredient_id}"
            if duplicate_key in seen_ingredient_keys:
                print(
                    f"   [AVISO] Ingrediente duplicado ignorado na receita "
                    f"'{recipe_name}': {row.ingredient_name}"
                )
                continue

            seen_ingredient_keys.add(duplicate_key)

            api.post(
                f"/recipes/{recipe_id}/ingredients",
                {
                    "ingredient_id": ingredient_id,
                    "quantity": row.quantity,
                    "unit": row.unit,
                },
            )

        updated_recipes += 1

    print(f"==> Ingredientes por receita: {updated_recipes} receitas sincronizadas")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Importa receitas familiares de um ficheiro Excel para o NutriFlow."
    )
    parser.add_argument(
        "--file",
        type=Path,
        default=DEFAULT_FILE,
        help=f"Caminho do ficheiro Excel (default: {DEFAULT_FILE})",
    )
    parser.add_argument(
        "--base-url",
        default=DEFAULT_BASE_URL,
        help=f"Base URL da API (default: {DEFAULT_BASE_URL})",
    )
    parser.add_argument(
        "--mode",
        choices=["merge", "reset"],
        default="merge",
        help="merge = atualiza/cria; reset = tenta apagar catálogo antes de importar",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    if not args.file.exists():
        print(f"ERRO: ficheiro não encontrado: {args.file}")
        return 1

    print(f"==> A abrir Excel: {args.file}")
    workbook = load_workbook(filename=args.file, data_only=True)

    recipes = parse_recipes(workbook)
    products = parse_products(workbook)
    recipe_ingredients = parse_recipe_ingredients(workbook)

    print(f"==> Receitas lidas: {len(recipes)}")
    print(f"==> Produtos lidos: {len(products)}")
    print(f"==> Linhas de ingredientes lidas: {len(recipe_ingredients)}")

    api = ApiClient(args.base_url)

    if args.mode == "reset":
        reset_catalog(api)

    ingredient_ids_by_key = upsert_ingredients(api, products)
    recipe_ids_by_key = upsert_recipes(api, recipes)
    grouped_ingredients = group_ingredients_by_recipe(recipe_ingredients)
    replace_recipe_ingredients(api, recipe_ids_by_key, ingredient_ids_by_key, grouped_ingredients)

    print("==> Importação concluída com sucesso")
    return 0


if __name__ == "__main__":
    sys.exit(main())